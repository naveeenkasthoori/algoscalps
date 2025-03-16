"""
Report generator for trading system results.
"""
import os
import json
import pandas as pd
import numpy as np
import logging
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any, Tuple
import matplotlib.pyplot as plt
import seaborn as sns

logger = logging.getLogger(__name__)

class ReportGenerator:
    """
    Generates performance reports for the trading system.
    """
    
    def __init__(self, output_dir: str, currency_symbol: str = "₹"):
        """
        Initialize the report generator.
        
        Args:
            output_dir: Directory for output files
            currency_symbol: Currency symbol to use in reports (default: ₹)
        """
        self.output_dir = output_dir
        self.currency_symbol = currency_symbol
        os.makedirs(output_dir, exist_ok=True)
        logger.info(f"Report generator initialized with output directory: {output_dir}")
    
    def _format_currency(self, value: float) -> str:
        """
        Format a value as currency with the appropriate symbol.
        
        Args:
            value: Value to format
            
        Returns:
            Formatted currency string
        """
        return f"{self.currency_symbol}{value:,.2f}"
    
    def _write_html_file(self, html_content: List[str], filepath: str) -> bool:
        """
        Write HTML content to file with proper Unicode handling.
        
        Args:
            html_content: List of HTML content lines
            filepath: Path to write the file
            
        Returns:
            True if successful, False otherwise
        """
        try:
            # Use UTF-8 encoding to handle Unicode characters
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write("\n".join(html_content))
            return True
        except Exception as e:
            logger.error(f"Error writing HTML file: {e}")
            
            # Fallback: Try to write without currency symbols
            try:
                modified_content = [line.replace(self.currency_symbol, '') for line in html_content]
                with open(filepath, 'w', encoding='utf-8') as f:
                    f.write("\n".join(modified_content))
                logger.warning(f"Wrote file without currency symbols: {filepath}")
                return True
            except Exception as fallback_error:
                logger.error(f"Fallback also failed: {fallback_error}")
                return False
    
    def generate_backtest_report(self, backtest_results: Dict[str, Any]) -> str:
        """
        Generate backtest report in Excel format.
        
        Args:
            backtest_results: Dictionary containing backtest results
            
        Returns:
            Path to Excel report
        """
        # Generate Excel report
        excel_path = self.generate_excel_report(backtest_results)
        
        # Add Excel report path to results
        backtest_results['excel_report_path'] = excel_path
        
        return excel_path
    
    def generate_html_report(self, backtest_results: Dict[str, Any]) -> str:
        """
        Generate HTML report for backtest results.
        
        This method is currently not implemented. Use generate_excel_report instead.
        
        Args:
            backtest_results: Dictionary containing backtest results
            
        Returns:
            Path to HTML report file (placeholder)
        """
        logger.warning("HTML report generation is not implemented. Using Excel report instead.")
        return "HTML_REPORT_NOT_IMPLEMENTED"
    
    def _get_exit_description(self, exit_reason: str) -> str:
        """
        Get a detailed description for an exit reason.
        
        Args:
            exit_reason: Exit reason code
            
        Returns:
            Detailed description
        """
        descriptions = {
            "stop_loss": "Price hit the stop loss level",
            "time_stop": "Maximum holding time for losing trade was reached",
            "signal_reversal": "Market signals reversed direction",
            "first_target": "Price reached first profit target",
            "second_target": "Price reached second profit target",
            "delta_reversal": "Option delta reversed direction significantly",
            "session_end": "End of trading session",
            "new_trading_day": "Position closed at end of trading day",
            "forced_exit": "Position forcibly closed (e.g., backtest end)",
            "pcr_reversal": "Put-Call Ratio reversed trend",
            "direction_reversal": "Market direction reversed"
        }
        
        return descriptions.get(exit_reason, exit_reason)
    
    def generate_trade_log_csv(self, trades: List[Dict[str, Any]]) -> str:
        """
        Generate a CSV file with detailed trade information.
        
        Args:
            trades: List of trade dictionaries
            
        Returns:
            Path to the generated CSV file
        """
        if not trades:
            logger.warning("No trades to export")
            return ""
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        csv_file = os.path.join(self.output_dir, f"trade_log_{timestamp}.csv")
        
        try:
            # Prepare data with enhanced information
            trade_data = []
            for trade in trades:
                # Extract basic trade information
                entry_time = trade.get('entry_time', '')
                exit_time = trade.get('exit_time', '')
                option_type = trade.get('option_type', '')
                strike_price = trade.get('strike_price', '')
                expiry_date = trade.get('expiry_date', '')
                entry_price = round(trade.get('entry_price', 0), 2)  # Round to 2 decimal places
                exit_price = round(trade.get('exit_price', 0), 2)    # Round to 2 decimal places
                position_size = trade.get('position_size', 0)
                
                # Format monetary values with 2 decimal places
                pnl = round(trade.get('total_net_pnl', trade.get('net_pnl', trade.get('pnl', 0))), 2)
                pnl_percent = round(trade.get('pnl_percent', 0) * 100, 2)  # Convert to percentage and round
                
                # Get entry conditions
                signal = trade.get('signal', {})
                conditions = signal.get('conditions', {}) if signal else trade.get('conditions', {})
                conditions_met = signal.get('conditions_met', 0) if signal else trade.get('conditions_met', 0)
                
                # Format conditions met as a comma-separated list
                met_conditions = []
                for cond_name, cond_met in conditions.items():
                    if cond_met:
                        met_conditions.append(cond_name)
                conditions_str = ", ".join(met_conditions)
                
                # Calculate and format trade duration
                duration_minutes = round(trade.get('duration_minutes', 0), 1)
                duration_formatted = f"{int(duration_minutes // 60)}h {int(duration_minutes % 60)}m" if duration_minutes >= 60 else f"{int(duration_minutes)}m"
                
                # Get detailed exit reason
                exit_reason = trade.get('exit_reason', '')
                exit_description = self._get_exit_description(exit_reason)
                
                # Extract transaction costs
                entry_costs = round(trade.get('entry_costs', {}).get('total', 0), 2)
                exit_costs = round(trade.get('exit_costs', {}).get('total', 0), 2)
                total_costs = round(entry_costs + exit_costs, 2)
                
                # Create enhanced trade record
                enhanced_trade = {
                    'trade_id': trade.get('id', ''),
                    'entry_time': entry_time,
                    'exit_time': exit_time,
                    'option_type': option_type,
                    'strike_price': strike_price,
                    'expiry_date': expiry_date,
                    'entry_price': entry_price,
                    'exit_price': exit_price,
                    'position_size': position_size,
                    'pnl': pnl,
                    'pnl_percent': pnl_percent,
                    'duration': duration_formatted,
                    'duration_minutes': duration_minutes,
                    'exit_reason': exit_reason,
                    'exit_description': exit_description,
                    'entry_conditions_met': conditions_str,
                    'conditions_count': conditions_met,
                    'entry_costs': entry_costs,
                    'exit_costs': exit_costs,
                    'total_costs': total_costs,
                    'delta': round(trade.get('delta', 0), 2),
                    'gamma': round(trade.get('gamma', 0), 4),
                    'original_size': trade.get('original_size', position_size),
                    'partial_exits': len(trade.get('partial_exits', [])),
                    'entry_signal_score': signal.get('quality_score', 0) if signal else 0,
                }
                
                trade_data.append(enhanced_trade)
            
            # Create DataFrame and save to CSV
            df = pd.DataFrame(trade_data)
            df.to_csv(csv_file, index=False, float_format='%.2f')  # Ensure all floats are formatted with 2 decimal places
            
            logger.info(f"Trade log CSV generated: {csv_file}")
            return csv_file
            
        except Exception as e:
            logger.error(f"Error generating trade log CSV: {e}")
            return ""
    
    def generate_daily_performance_report(self, daily_pnl: Dict[datetime.date, float]) -> str:
        """
        Generate a report of daily performance.
        
        Args:
            daily_pnl: Dictionary with daily P&L by date
            
        Returns:
            Path to the generated report file
        """
        if not daily_pnl:
            logger.warning("No daily P&L data to report")
            # Generate a simple report even when no data is available
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            report_file = os.path.join(self.output_dir, f"daily_performance_{timestamp}.html")
            
            html_content = []
            html_content.append("<!DOCTYPE html>")
            html_content.append("<html><head>")
            html_content.append("<meta charset='utf-8'>")
            html_content.append("<title>Daily Performance Report</title>")
            html_content.append("<style>body { font-family: Arial, sans-serif; margin: 20px; }</style>")
            html_content.append("</head><body>")
            html_content.append("<h1>Daily Performance Report</h1>")
            html_content.append("<p>No daily performance data available for the selected period.</p>")
            html_content.append("</body></html>")
            
            if self._write_html_file(html_content, report_file):
                return report_file
            return ""
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_file = os.path.join(self.output_dir, f"daily_performance_{timestamp}.html")
        
        try:
            # Convert to DataFrame
            df = pd.Series(daily_pnl).reset_index()
            df.columns = ['Date', 'PnL']
            df = df.sort_values('Date')
            
            # Calculate cumulative P&L
            df['Cumulative'] = df['PnL'].cumsum()
            
            # Generate chart
            chart_file = os.path.join(self.output_dir, f"daily_pnl_{timestamp}.png")
            plt.figure(figsize=(12, 8))
            
            plt.subplot(2, 1, 1)
            colors = ['green' if x > 0 else 'red' for x in df['PnL']]
            plt.bar(df['Date'], df['PnL'], color=colors)
            plt.title('Daily P&L')
            plt.ylabel(f'P&L ({self.currency_symbol})')
            plt.grid(True, axis='y')
            
            plt.subplot(2, 1, 2)
            plt.plot(df['Date'], df['Cumulative'], 'b-')
            plt.title('Cumulative P&L')
            plt.ylabel(f'Cumulative P&L ({self.currency_symbol})')
            plt.grid(True)
            
            plt.tight_layout()
            plt.savefig(chart_file)
            plt.close()
            
            # Generate HTML report
            html_content = []
            html_content.append("<!DOCTYPE html>")
            html_content.append("<html><head>")
            html_content.append("<meta charset='utf-8'>")
            html_content.append("<style>")
            html_content.append("body { font-family: Arial, sans-serif; margin: 20px; }")
            html_content.append("h1, h2 { color: #333366; }")
            html_content.append("table { border-collapse: collapse; width: 100%; margin-bottom: 20px; }")
            html_content.append("th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }")
            html_content.append("th { background-color: #f2f2f2; }")
            html_content.append("tr:nth-child(even) { background-color: #f9f9f9; }")
            html_content.append(".positive { color: green; }")
            html_content.append(".negative { color: red; }")
            html_content.append("</style>")
            html_content.append("<title>Daily Performance Report</title>")
            html_content.append("</head><body>")
            
            html_content.append("<h1>Daily Performance Report</h1>")
            html_content.append(f"<p>Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>")
            
            html_content.append("<h2>Performance Chart</h2>")
            html_content.append(f"<img src='{os.path.basename(chart_file)}' width='800'>")
            
            html_content.append("<h2>Daily P&L Table</h2>")
            html_content.append("<table>")
            html_content.append("<tr><th>Date</th><th>P&L</th><th>Cumulative P&L</th></tr>")
            
            for _, row in df.iterrows():
                date = row['Date']
                pnl = row['PnL']
                cumulative = row['Cumulative']
                
                pnl_class = "positive" if pnl > 0 else "negative"
                cumulative_class = "positive" if cumulative > 0 else "negative"
                
                html_content.append("<tr>")
                html_content.append(f"<td>{date}</td>")
                html_content.append(f"<td class='{pnl_class}'>{self._format_currency(pnl)}</td>")
                html_content.append(f"<td class='{cumulative_class}'>{self._format_currency(cumulative)}</td>")
                html_content.append("</tr>")
            
            html_content.append("</table>")
            
            # Summary statistics
            positive_days = sum(1 for pnl in daily_pnl.values() if pnl > 0)
            negative_days = sum(1 for pnl in daily_pnl.values() if pnl <= 0)
            total_days = len(daily_pnl)
            win_rate = positive_days / total_days * 100 if total_days > 0 else 0
            
            avg_win = np.mean([pnl for pnl in daily_pnl.values() if pnl > 0]) if positive_days > 0 else 0
            avg_loss = abs(np.mean([pnl for pnl in daily_pnl.values() if pnl <= 0])) if negative_days > 0 else 0
            
            html_content.append("<h2>Summary Statistics</h2>")
            html_content.append("<table>")
            html_content.append("<tr><th>Metric</th><th>Value</th></tr>")
            html_content.append(f"<tr><td>Total Trading Days</td><td>{total_days}</td></tr>")
            html_content.append(f"<tr><td>Profitable Days</td><td>{positive_days}</td></tr>")
            html_content.append(f"<tr><td>Losing Days</td><td>{negative_days}</td></tr>")
            html_content.append(f"<tr><td>Win Rate</td><td>{win_rate:.1f}%</td></tr>")
            html_content.append(f"<tr><td>Average Daily Profit</td><td class='positive'>{self._format_currency(avg_win)}</td></tr>")
            html_content.append(f"<tr><td>Average Daily Loss</td><td class='negative'>{self._format_currency(avg_loss)}</td></tr>")
            html_content.append("</table>")
            
            html_content.append("</body></html>")
            
            # Write HTML to file with proper encoding
            if self._write_html_file(html_content, report_file):
                logger.info(f"Daily performance report generated: {report_file}")
                return report_file
            else:
                logger.error("Failed to write daily performance report")
                return ""
            
        except Exception as e:
            logger.error(f"Error generating daily performance report: {e}")
            return ""
    
    def generate_debug_report(self, 
                            spot_data: pd.DataFrame, 
                            futures_data: pd.DataFrame,
                            options_data: Dict[str, pd.DataFrame],
                            config: Dict[str, Any]) -> str:
        """
        Generate a detailed debug report for analyzing data issues.
        
        Args:
            spot_data: Spot market data
            futures_data: Futures market data
            options_data: Dictionary of options data by expiry
            config: Backtest configuration parameters
            
        Returns:
            Path to the generated debug report file
        """
        logger.info("Generating data debug report")
        
        # Create timestamp for filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_file = os.path.join(self.output_dir, f"debug_report_{timestamp}.html")
        
        # Generate HTML report
        html_content = []
        html_content.append("<!DOCTYPE html>")
        html_content.append("<html><head>")
        html_content.append("<meta charset='utf-8'>")
        html_content.append("<style>")
        html_content.append("body { font-family: Arial, sans-serif; margin: 20px; }")
        html_content.append("h1, h2, h3 { color: #333366; }")
        html_content.append("table { border-collapse: collapse; width: 100%; margin-bottom: 20px; }")
        html_content.append("th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }")
        html_content.append("th { background-color: #f2f2f2; }")
        html_content.append("tr:nth-child(even) { background-color: #f9f9f9; }")
        html_content.append(".error { color: red; font-weight: bold; }")
        html_content.append(".warning { color: orange; }")
        html_content.append(".success { color: green; }")
        html_content.append(".code { font-family: monospace; background-color: #f5f5f5; padding: 10px; border-radius: 5px; overflow-x: auto; }")
        html_content.append("</style>")
        html_content.append("<title>Backtest Data Debug Report</title>")
        html_content.append("</head><body>")
        
        # Header
        html_content.append(f"<h1>Backtest Data Debug Report</h1>")
        html_content.append(f"<p>Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>")
        
        # Configuration
        html_content.append("<h2>Backtest Configuration</h2>")
        html_content.append("<table>")
        html_content.append("<tr><th>Parameter</th><th>Value</th></tr>")
        for key, value in config.items():
            html_content.append(f"<tr><td>{key}</td><td>{value}</td></tr>")
        html_content.append("</table>")
        
        # Spot Data Summary
        html_content.append("<h2>Spot Data Summary</h2>")
        if spot_data.empty:
            html_content.append("<p class='error'>No spot data available!</p>")
        else:
            date_range = f"{spot_data.index.min()} to {spot_data.index.max()}"
            html_content.append(f"<p><strong>Date Range:</strong> {date_range}</p>")
            html_content.append(f"<p><strong>Number of Rows:</strong> {len(spot_data)}</p>")
            html_content.append("<p><strong>Columns:</strong></p>")
            html_content.append("<ul>")
            for col in spot_data.columns:
                html_content.append(f"<li>{col}</li>")
            html_content.append("</ul>")
            
            # Sample data
            html_content.append("<h3>Sample Data (First 5 Rows)</h3>")
            html_content.append("<div class='code'>")
            html_content.append(spot_data.head(5).to_html())
            html_content.append("</div>")
        
        # Futures Data Summary
        html_content.append("<h2>Futures Data Summary</h2>")
        if futures_data.empty:
            html_content.append("<p class='error'>No futures data available!</p>")
        else:
            date_range = f"{futures_data.index.min()} to {futures_data.index.max()}"
            html_content.append(f"<p><strong>Date Range:</strong> {date_range}</p>")
            html_content.append(f"<p><strong>Number of Rows:</strong> {len(futures_data)}</p>")
            html_content.append("<p><strong>Columns:</strong></p>")
            html_content.append("<ul>")
            for col in futures_data.columns:
                html_content.append(f"<li>{col}</li>")
            html_content.append("</ul>")
            
            # Sample data
            html_content.append("<h3>Sample Data (First 5 Rows)</h3>")
            html_content.append("<div class='code'>")
            html_content.append(futures_data.head(5).to_html())
            html_content.append("</div>")
        
        # Options Data Summary
        html_content.append("<h2>Options Data Summary</h2>")
        if not options_data:
            html_content.append("<p class='error'>No options data available!</p>")
        else:
            html_content.append(f"<p><strong>Number of Expiry Dates:</strong> {len(options_data)}</p>")
            
            html_content.append("<table>")
            html_content.append("<tr><th>Expiry Date</th><th>Date Range</th><th>Rows</th><th>Strike Range</th><th>Status</th></tr>")
            
            # Check if any options expire after the backtest start date
            start_date = pd.Timestamp(config.get('start_date', '1900-01-01'))
            
            for expiry, df in options_data.items():
                try:
                    expiry_date = pd.to_datetime(expiry)
                    date_range = f"{df.index.min()} to {df.index.max()}"
                    num_rows = len(df)
                    
                    if 'strike_price' in df.columns:
                        strike_range = f"{df['strike_price'].min()} to {df['strike_price'].max()}"
                    else:
                        strike_range = "N/A"
                    
                    status_class = "success" if expiry_date >= start_date else "error"
                    status_text = "Valid" if expiry_date >= start_date else "Expired before backtest"
                    
                    html_content.append(f"<tr>")
                    html_content.append(f"<td>{expiry}</td>")
                    html_content.append(f"<td>{date_range}</td>")
                    html_content.append(f"<td>{num_rows}</td>")
                    html_content.append(f"<td>{strike_range}</td>")
                    html_content.append(f"<td class='{status_class}'>{status_text}</td>")
                    html_content.append(f"</tr>")
                    
                except Exception as e:
                    html_content.append(f"<tr>")
                    html_content.append(f"<td>{expiry}</td>")
                    html_content.append(f"<td colspan='3' class='error'>Error: {str(e)}</td>")
                    html_content.append(f"<td class='error'>Invalid</td>")
                    html_content.append(f"</tr>")
            
            html_content.append("</table>")
        
        # Date Compatibility Analysis
        html_content.append("<h2>Date Compatibility Analysis</h2>")
        
        # Check for overlap between backtest dates and available data
        start_date = pd.Timestamp(config.get('start_date', '1900-01-01'))
        end_date = pd.Timestamp(config.get('end_date', '2100-01-01'))
        
        spot_start = spot_data.index.min() if not spot_data.empty else None
        spot_end = spot_data.index.max() if not spot_data.empty else None
        
        futures_start = futures_data.index.min() if not futures_data.empty else None
        futures_end = futures_data.index.max() if not futures_data.empty else None
        
        html_content.append("<table>")
        html_content.append("<tr><th>Data Source</th><th>Start Date</th><th>End Date</th><th>Overlaps with Backtest</th></tr>")
        
        # Spot data
        spot_overlaps = (spot_start <= end_date and spot_end >= start_date) if spot_start and spot_end else False
        spot_class = "success" if spot_overlaps else "error"
        html_content.append(f"<tr><td>Spot Data</td><td>{spot_start}</td><td>{spot_end}</td><td class='{spot_class}'>{spot_overlaps}</td></tr>")
        
        # Futures data
        futures_overlaps = (futures_start <= end_date and futures_end >= start_date) if futures_start and futures_end else False
        futures_class = "success" if futures_overlaps else "error"
        html_content.append(f"<tr><td>Futures Data</td><td>{futures_start}</td><td>{futures_end}</td><td class='{futures_class}'>{futures_overlaps}</td></tr>")
        
        # Options data (by expiry)
        any_options_overlap = False
        for expiry, df in options_data.items():
            options_start = df.index.min()
            options_end = df.index.max()
            
            try:
                expiry_date = pd.to_datetime(expiry)
                options_overlaps = (options_start <= end_date and options_end >= start_date and expiry_date >= start_date)
                any_options_overlap = any_options_overlap or options_overlaps
                options_class = "success" if options_overlaps else "error"
                
                html_content.append(f"<tr><td>Options (Expiry: {expiry})</td><td>{options_start}</td><td>{options_end}</td><td class='{options_class}'>{options_overlaps}</td></tr>")
            except:
                html_content.append(f"<tr><td>Options (Expiry: {expiry})</td><td>{options_start}</td><td>{options_end}</td><td class='error'>Invalid expiry format</td></tr>")
        
        html_content.append("</table>")
        
        # Overall compatibility assessment
        overall_compatible = spot_overlaps and futures_overlaps and any_options_overlap
        compatibility_class = "success" if overall_compatible else "error"
        compatibility_message = "All data sources overlap with backtest period" if overall_compatible else "Data sources do not properly overlap with backtest period"
        
        html_content.append(f"<h3>Overall Compatibility: <span class='{compatibility_class}'>{compatibility_message}</span></h3>")
        
        # Recommendations
        html_content.append("<h2>Recommendations</h2>")
        html_content.append("<ul>")
        
        if not overall_compatible:
            if not spot_overlaps:
                html_content.append("<li class='error'>Spot data does not cover the backtest period. Consider adjusting the backtest date range or providing spot data for the required period.</li>")
            
            if not futures_overlaps:
                html_content.append("<li class='error'>Futures data does not cover the backtest period. Consider adjusting the backtest date range or providing futures data for the required period.</li>")
            
            if not any_options_overlap:
                html_content.append("<li class='error'>No options data overlaps with the backtest period. Either:</li>")
                html_content.append("<ul>")
                html_content.append("<li>Adjust the backtest date range to match available options data</li>")
                html_content.append("<li>Provide options data that includes valid expiry dates within or after the backtest period</li>")
                html_content.append("<li>Check the expiry date formats in the options data files</li>")
                html_content.append("</ul>")
        
        # Suggest using auto-dates feature
        html_content.append("<li>Consider using the <code>--auto-dates</code> flag to automatically select a compatible date range based on the available data.</li>")
        
        # Check for expiry date format issues
        expiry_format_issues = False
        for expiry in options_data.keys():
            try:
                pd.to_datetime(expiry)
            except:
                expiry_format_issues = True
                break
        
        if expiry_format_issues:
            html_content.append("<li class='error'>Some expiry dates have invalid formats. Ensure all expiry dates are in a valid datetime format (e.g., YYYY-MM-DD).</li>")
        
        html_content.append("</ul>")
        
        # Footer
        html_content.append("<hr>")
        html_content.append("<p><i>Generated by Options Trading System - Debug Report</i></p>")
        html_content.append("</body></html>")
        
        # Write HTML to file with proper encoding
        if self._write_html_file(html_content, report_file):
            logger.info(f"Debug report generated: {report_file}")
            return report_file
        else:
            logger.error("Failed to write debug report")
            return ""
    
    def generate_excel_report(self, backtest_results: Dict[str, Any]) -> str:
        """
        Generate detailed Excel report with performance metrics.
        
        Args:
            backtest_results: Dictionary containing backtest results
            
        Returns:
            Path to saved Excel file
        """
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font
            from openpyxl.utils import get_column_letter
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_path = os.path.join(self.output_dir, f"backtest_report_{timestamp}.xlsx")
            
            # Create Excel writer object
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                # 1. Summary Sheet
                summary_data = {
                    'Metric': [
                        'Total Trades',
                        'Winning Trades',
                        'Losing Trades',
                        'Win Rate',
                        'Total P&L',
                        'Total Profit',
                        'Total Loss',
                        'Profit Factor',
                        'Max Drawdown',
                        'Total Costs',
                        'Net P&L',
                    ],
                    'Value': [
                        backtest_results['num_trades'],
                        backtest_results['winning_trades'],
                        backtest_results['losing_trades'],
                        f"{backtest_results['win_rate']*100:.1f}%",
                        f"₹{backtest_results['total_pnl']:,.2f}",
                        f"₹{backtest_results['total_profit']:,.2f}",
                        f"₹{backtest_results['total_loss']:,.2f}",
                        f"{backtest_results['profit_factor']:.2f}",
                        f"{backtest_results['max_drawdown']*100:.1f}%",
                        f"₹{backtest_results['total_costs']:,.2f}",
                        f"₹{backtest_results['total_pnl']-backtest_results['total_costs']:,.2f}",
                    ]
                }
                
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                
                # 2. Daily Performance Sheet
                if 'daily_pnl' in backtest_results:
                    daily_df = pd.DataFrame({
                        'Date': backtest_results['daily_pnl'].keys(),
                        'P&L': backtest_results['daily_pnl'].values()
                    })
                    daily_df['Cumulative P&L'] = daily_df['P&L'].cumsum()
                    
                    # Calculate daily drawdown
                    daily_df['Peak'] = daily_df['Cumulative P&L'].cummax()
                    daily_df['Drawdown'] = (daily_df['Cumulative P&L'] - daily_df['Peak']) / daily_df['Peak'] * 100
                    
                    daily_df.to_excel(writer, sheet_name='Daily Performance', index=False)
                
                # 3. Trade Log Sheet
                if 'trades' in backtest_results:
                    trades_data = []
                    for trade in backtest_results['trades']:
                        trade_data = {
                            'Entry Time': trade['entry_time'],
                            'Exit Time': trade['exit_time'],
                            'Type': trade['option_type'],
                            'Strike': trade['strike_price'],
                            'Entry Price': trade['entry_price'],
                            'Exit Price': trade['exit_price'],
                            'Size': trade['position_size'],
                            'P&L': trade['pnl'],
                            'Net P&L': trade['net_pnl'],
                            'Costs': trade['all_costs'],
                            'Duration (min)': trade['duration_minutes'],
                            'Exit Reason': trade['exit_reason']
                        }
                        trades_data.append(trade_data)
                    
                    trades_df = pd.DataFrame(trades_data)
                    trades_df.to_excel(writer, sheet_name='Trade Log', index=False)
                
                # 4. Drawdown Analysis Sheet
                if 'equity_curve' in backtest_results:
                    equity_curve = backtest_results['equity_curve']
                    running_max = equity_curve.cummax()
                    drawdown = (equity_curve - running_max) / running_max * 100
                    
                    drawdown_df = pd.DataFrame({
                        'Date': equity_curve.index,
                        'Equity': equity_curve.values,
                        'Peak': running_max.values,
                        'Drawdown %': drawdown.values
                    })
                    
                    # Find worst drawdowns
                    worst_drawdowns = drawdown_df.nsmallest(10, 'Drawdown %')
                    
                    drawdown_df.to_excel(writer, sheet_name='Drawdown Analysis', index=False)
                    worst_drawdowns.to_excel(writer, sheet_name='Worst Drawdowns', index=False)
                
                # 5. Market Regime Analysis
                if 'regime_statistics' in backtest_results:
                    regime_data = []
                    for regime, stats in backtest_results['regime_statistics'].items():
                        regime_data.append({
                            'Regime': regime,
                            'Trade Count': stats['count'],
                            'Win Rate': f"{stats['win_rate']*100:.1f}%",
                            'Total P&L': f"₹{stats['total_pnl']:,.2f}"
                        })
                    
                    regime_df = pd.DataFrame(regime_data)
                    regime_df.to_excel(writer, sheet_name='Regime Analysis', index=False)
                
                # Auto-adjust column widths
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    for column in worksheet.columns:
                        max_length = 0
                        column = [cell for cell in column]
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = (max_length + 2)
                        worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
            
            logger.info(f"Excel report generated: {excel_path}")
            return excel_path
            
        except Exception as e:
            logger.error(f"Error generating Excel report: {e}")
            return ""

